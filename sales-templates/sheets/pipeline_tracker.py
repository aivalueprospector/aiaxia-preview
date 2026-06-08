#!/usr/bin/env python3
"""
Branded Prospect / Pipeline tracker (Google Sheets via .xlsx).
Lifecycle stages from twincloud/docs/customer-lifecycle-terminology.md (Prospect → Pro → Member).
Est. value uses the Rate Card (PRICING-SPEC.md). One shared AIVP tracker covering both brands.

Usage:  python sheets/pipeline_tracker.py
Output: en/aivp-pipeline-tracker.xlsx
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent.parent
REPO = HERE.parent
BRANDS = json.loads((HERE / "brands.json").read_text())
B = BRANDS["aivp"]

TIERS = ["SILVER", "GOLD", "PLATINUM", "DIAMOND"]
STAGES = ["Prospect", "Qualified", "Demo Scheduled", "Proposal Sent",
          "Pro (Won)", "Member", "On Hold", "Lost"]
RATE_ANNUAL = {"SILVER": 199, "GOLD": 349, "PLATINUM": 489, "DIAMOND": 689}

THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(THIN, THIN, THIN, THIN)
INPUT_FILL = PatternFill("solid", fgColor="FFFDF5")
ROWS = 100


def build():
    accent, accent_dark = B["accent"], B["accent_dark"]
    wb = openpyxl.Workbook()

    # ---- Rate Card (hidden helper for VLOOKUP) ----
    rc = wb.create_sheet("Rate Card")
    rc.append(["Tier", "Annual Fee"])
    for t in TIERS:
        rc.append([t, RATE_ANNUAL[t]])
    rc.sheet_state = "hidden"

    # ---- Pipeline ----
    ws = wb.active
    ws.title = "Pipeline"
    ws.sheet_view.showGridLines = False
    headers = ["Company / Prospect", "Contact", "Email", "Brand", "Stage",
               "Tier", "Seats", "Est. Annual Value", "Next Action",
               "Next Action Date", "Notes"]
    widths = [26, 18, 26, 12, 16, 12, 8, 16, 26, 16, 30]

    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "AI Value Prospectors — Sales Pipeline Tracker"
    t.fill = PatternFill("solid", fgColor=accent)
    t.font = Font(color="0F1D2C", bold=True, size=15)
    t.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.append(headers)
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill = PatternFill("solid", fgColor=accent_dark)
        c.font = Font(color="FFFFFF", bold=True)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[chr(64 + i)].width = w

    # example row
    example = ["Acme Logistics", "Maria Vega", "maria@acme.com", "ProAsiste",
               "Demo Scheduled", "GOLD", 5, None, "Send recap + quote", "", "Warm intro via referral"]
    ws.append(example)

    # data rows with validations + Est. value formula
    brand_dv = DataValidation(type="list", formula1='"ProAsiste,EduAsiste"', allow_blank=True)
    stage_dv = DataValidation(type="list", formula1=f'"{",".join(STAGES)}"', allow_blank=True)
    tier_dv = DataValidation(type="list", formula1=f'"{",".join(TIERS)}"', allow_blank=True)
    date_dv = DataValidation(type="date", allow_blank=True)
    for dv in (brand_dv, stage_dv, tier_dv, date_dv):
        ws.add_data_validation(dv)

    first, last = 3, 2 + ROWS
    for r in range(first, last + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            if col in (1, 2, 3, 7, 9, 11):
                cell.fill = INPUT_FILL
        brand_dv.add(ws.cell(row=r, column=4))
        stage_dv.add(ws.cell(row=r, column=5))
        tier_dv.add(ws.cell(row=r, column=6))
        date_dv.add(ws.cell(row=r, column=10))
        # Est annual value = seats * annual rate (blank when tier/seats empty)
        ws.cell(row=r, column=8).value = (
            f'=IF(OR(F{r}="",G{r}=""),"",G{r}*VLOOKUP(F{r},'
            f"'Rate Card'!$A$2:$B$5,2,FALSE))"
        )
        ws.cell(row=r, column=8).number_format = '$#,##0'

    ws.freeze_panes = "A3"

    # ---- Summary ----
    sm = wb.create_sheet("Summary")
    sm.sheet_view.showGridLines = False
    sm.column_dimensions["A"].width = 20
    sm.column_dimensions["B"].width = 14
    sm.column_dimensions["C"].width = 18
    sm.merge_cells("A1:C1")
    sm["A1"] = "Pipeline Summary"
    sm["A1"].fill = PatternFill("solid", fgColor=accent)
    sm["A1"].font = Font(color="0F1D2C", bold=True, size=14)
    sm.append(["Stage", "Count", "Est. Annual Value"])
    for c in sm[2]:
        c.fill = PatternFill("solid", fgColor=accent_dark); c.font = Font(color="FFFFFF", bold=True); c.border = BORDER
    rng_stage = f"Pipeline!$E${first}:$E${last}"
    rng_val = f"Pipeline!$H${first}:$H${last}"
    for st in STAGES:
        sm.append([
            st,
            f'=COUNTIF({rng_stage},"{st}")',
            f'=SUMIF({rng_stage},"{st}",{rng_val})',
        ])
    for row in sm.iter_rows(min_row=3, max_row=2 + len(STAGES), max_col=3):
        for c in row:
            c.border = BORDER
            if c.column == 3:
                c.number_format = '$#,##0'
    tot = 3 + len(STAGES)
    sm.cell(row=tot, column=1, value="TOTAL").font = Font(bold=True)
    sm.cell(row=tot, column=2, value=f"=SUM(B3:B{tot-1})").font = Font(bold=True)
    sm.cell(row=tot, column=3, value=f"=SUM(C3:C{tot-1})").font = Font(bold=True)
    sm.cell(row=tot, column=3).number_format = '$#,##0'

    out = HERE / "en" / "aivp-pipeline-tracker.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(str(out))
    return out


if __name__ == "__main__":
    print(f"  built {build().name}")
