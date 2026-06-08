#!/usr/bin/env python3
"""
Branded Pricing + ROI calculator (Google Sheets via .xlsx).
Rate card + certification-waiver logic from twincloud/docs/PRICING-SPEC.md (Model E, v1.1).

Usage:  python sheets/roi_calculator.py [brand ...]   (default: all brands)
Outputs: en/{brand}-pricing-roi-calculator.xlsx
Live formulas recompute in Google Sheets after import. Input cells are yellow.
"""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent.parent
REPO = HERE.parent
BRANDS = json.loads((HERE / "brands.json").read_text())

# Rate card — PRICING-SPEC.md §1.1 / §1.3
TIERS = ["SILVER", "GOLD", "PLATINUM", "DIAMOND"]
RATE = {  # tier: (setup, monthly, annual, addon, included_projects, min_seats)
    "SILVER":   (97, 20, 199, 12.99, 1, 1),
    "GOLD":     (332, 35, 349, 9.99, 3, 3),
    "PLATINUM": (997, 49, 489, 6.99, 6, 10),
    "DIAMOND":  (997, 69, 689, 4.99, 9, 50),
}
CERT_DEFAULT = 2750  # midpoint of $2,500–3,000 (TBD)

INPUT_FILL = PatternFill("solid", fgColor="FFF3B0")
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(THIN, THIN, THIN, THIN)
WHITE = Font(color="FFFFFF", bold=True)
MUTED = Font(color="6B7A8A", size=9, italic=True)


def build(brand_key, b):
    accent = b["accent"]
    accent_dark = b["accent_dark"]
    wb = openpyxl.Workbook()

    # ===== Rate Card sheet (reference for VLOOKUPs) =====
    rc = wb.active
    rc.title = "Rate Card"
    rc.append(["Tier", "Setup Fee", "Monthly Fee", "Annual Fee",
               "Add-on / project", "Included projects", "Min seats"])
    for i, t in enumerate(TIERS):
        rc.append([t, *RATE[t]])
    for cell in rc[1]:
        cell.fill = PatternFill("solid", fgColor=accent_dark)
        cell.font = WHITE
        cell.border = BORDER
    for row in rc.iter_rows(min_row=2, max_row=5, max_col=7):
        for c in row:
            c.border = BORDER
            if c.column > 1:
                c.number_format = '$#,##0.00'
    rc.append([])
    rc.append(["Certification course / person", CERT_DEFAULT, "PRICING-SPEC §2 (TBD $2,500–3,000)"])
    rc.append(["Annual discount", "~17%", "10 months for 12"])
    for col, w in zip("ABCDEFG", (28, 14, 14, 12, 16, 16, 12)):
        rc.column_dimensions[col].width = w

    # ===== Calculator sheet =====
    ws = wb.create_sheet("ROI Calculator", 0)
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", (34, 16, 16, 4, 40)):
        ws.column_dimensions[col].width = w

    # title band
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"{b['display']} — Pricing & ROI Calculator"
    t.fill = PatternFill("solid", fgColor=accent)
    t.font = Font(color="0F1D2C", bold=True, size=16)
    t.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    logo = REPO / b["logo"]
    if logo.exists():
        try:
            from openpyxl.drawing.image import Image as XImage
            img = XImage(str(logo)); img.height = 34; img.width = 34
            ws.add_image(img, "E1")
        except Exception:
            pass

    ws["A2"] = "Yellow cells are inputs. Everything else recalculates automatically."
    ws["A2"].font = MUTED

    # ---- inputs ----
    r = 4
    ws[f"A{r}"] = "INPUTS"; ws[f"A{r}"].font = Font(bold=True, color=accent_dark, size=12)
    inputs = [
        ("Tier", "GOLD", "list", ",".join(TIERS)),
        ("Number of seats", 5, "whole", None),
        ("Billing", "Annual", "list", "Monthly,Annual"),
        ("Add-on projects (total)", 0, "whole", None),
        ("Certified staff (setup waiver if ≥ 2)", 0, "whole", None),
        ("Certification course / person ($)", CERT_DEFAULT, None, None),
    ]
    input_cells = {}
    for i, (label, default, kind, formula) in enumerate(inputs):
        rr = r + 1 + i
        ws[f"A{rr}"] = label
        c = ws[f"B{rr}"]
        c.value = default
        c.fill = INPUT_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        input_cells[label.split(" (")[0]] = f"B{rr}"
        if kind == "list":
            dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=False)
            ws.add_data_validation(dv); dv.add(c)
        elif kind == "whole":
            dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0")
            ws.add_data_validation(dv); dv.add(c)
    tier = input_cells["Tier"]
    seats = input_cells["Number of seats"]
    billing = input_cells["Billing"]
    addons = input_cells["Add-on projects"]
    certified = input_cells["Certified staff"]
    course = input_cells["Certification course / person"]

    # ---- computed ----
    cr = r + len(inputs) + 2
    ws[f"A{cr}"] = "RESULTS"; ws[f"A{cr}"].font = Font(bold=True, color=accent_dark, size=12)
    RC = "'Rate Card'"
    setup_l = f"VLOOKUP({tier},{RC}!$A$2:$F$5,2,FALSE)"
    monthly_l = f"VLOOKUP({tier},{RC}!$A$2:$F$5,3,FALSE)"
    annual_l = f"VLOOKUP({tier},{RC}!$A$2:$F$5,4,FALSE)"
    addon_l = f"VLOOKUP({tier},{RC}!$A$2:$F$5,5,FALSE)"
    incl_l = f"VLOOKUP({tier},{RC}!$A$2:$F$5,6,FALSE)"

    # per-seat recurring depends on billing choice
    per_seat_year = f"IF({billing}=\"Annual\",{annual_l},{monthly_l}*12)"
    extra_proj = f"MAX(0,{addons}-{incl_l})"

    results = [
        ("Setup fee per seat", f"={setup_l}", '$#,##0.00'),
        ("Setup — without certification", f"={seats}*{setup_l}", '$#,##0.00'),
        ("Setup — with certification waiver", f"=IF({certified}>=2,{certified}*{setup_l},{seats}*{setup_l})", '$#,##0.00'),
        ("Certification course cost", f"={certified}*{course}", '$#,##0.00'),
        ("Recurring per seat / year", f"={per_seat_year}", '$#,##0.00'),
        ("Recurring — all seats / year", f"={seats}*{per_seat_year}", '$#,##0.00'),
        ("Add-on projects / year (all seats)", f"={seats}*{extra_proj}*{addon_l}*12", '$#,##0.00'),
        ("YEAR 1 TOTAL", f"=IF({certified}>=2,{certified}*{setup_l},{seats}*{setup_l})+{certified}*{course}+{seats}*{per_seat_year}+{seats}*{extra_proj}*{addon_l}*12", '$#,##0.00'),
        ("ONGOING / YEAR (yr 2+)", f"={seats}*{per_seat_year}+{seats}*{extra_proj}*{addon_l}*12", '$#,##0.00'),
        ("Setup saved by certification", f"=({seats}*{setup_l})-IF({certified}>=2,{certified}*{setup_l},{seats}*{setup_l})", '$#,##0.00'),
    ]
    for i, (label, formula, fmt) in enumerate(results):
        rr = cr + 1 + i
        ws[f"A{rr}"] = label
        c = ws[f"B{rr}"]; c.value = formula; c.number_format = fmt; c.border = BORDER
        if "TOTAL" in label or "ONGOING" in label:
            ws[f"A{rr}"].font = Font(bold=True, color="0F1D2C")
            c.font = Font(bold=True, color=accent_dark)
            ws[f"A{rr}"].fill = PatternFill("solid", fgColor=b["soft"])
            c.fill = PatternFill("solid", fgColor=b["soft"])

    note_r = cr + len(results) + 2
    ws[f"A{note_r}"] = ("Certification waiver (PRICING-SPEC §2.3): with ≥ 2 certified staff, "
                        "certified staff configure colleagues and colleague setup fees are $0. "
                        "Figures are estimates — confirm final pricing with AIVP before quoting.")
    ws[f"A{note_r}"].font = MUTED
    ws.merge_cells(f"A{note_r}:E{note_r}")
    ws[f"A{note_r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_r].height = 42

    ws.freeze_panes = "A4"
    out = HERE / "en" / f"{brand_key}-pricing-roi-calculator.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(str(out))
    return out


def main():
    which = [a for a in sys.argv[1:] if not a.startswith("--")] or list(BRANDS)
    for k in which:
        out = build(k, BRANDS[k])
        print(f"  built {out.name}")


if __name__ == "__main__":
    main()
